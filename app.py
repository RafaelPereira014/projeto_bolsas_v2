from datetime import date
import os
from flask import abort, request, redirect, url_for, flash
from flask import Flask, flash, redirect, request, jsonify, render_template, send_from_directory, session, url_for
from flask import request, send_file
from flask import render_template, request, send_file
import io
import csv
import requests
from openpyxl import Workbook
import pymysql
from werkzeug.security import generate_password_hash
from werkzeug.security import check_password_hash
from config import db_config
from db_operations.notifications import send_email_on_selection
from db_operations.selection.db_selection import *
from db_operations.consulting.db_consulting import *
from db_operations.user.db_user import *
from db_operations.admin.admin import *


app = Flask(__name__)
app.secret_key = 'bolsas_ilha'

# Set the upload folder
UPLOAD_FOLDER = '/Users/rafaelpereira/Desktop/projeto_Bolsas/static/ulpoads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {'csv'}
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def is_logged_in():
    # Check if the user_id exists in the session
    return 'user_id' in session

# Função para executar a consulta SQL
def execute_query(query, params):
    conn = create_connection()  # Criar a conexão
    cursor = conn.cursor(pymysql.cursors.DictCursor)  # Para retornar resultados como dicionário
    
    try:
        cursor.execute(query, params)  # Executar a query com os parâmetros
        results = cursor.fetchall()  # Obter todos os resultados da consulta
        return results
    except pymysql.Error as err:
        print(f"Erro: {err}")
        return None
    finally:
        cursor.close()  # Fechar o cursor
        conn.close()  # Fechar a conexão

def create_connection():
    connection = None
    try:
        
        connection = pymysql.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database=db_config['db']
        )
        
    except pymysql.Error as e:
        print(f"The error '{e}' occurred")
    return connection

# Route for rendering the frontend HTML
@app.route('/')
def index():
    return render_template('index.html')

# Login Route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        conn = create_connection()
        cursor = conn.cursor(pymysql.cursors.DictCursor)
        
        # Query to find the user
        query = "SELECT * FROM Admin WHERE email = %s"
        cursor.execute(query, (email,))
        user = cursor.fetchone()

        if user and check_password_hash(user['password'], password):
            # Password matches, log the user in
            session['user_id'] = user['id']
            session['email'] = user['email']
            session['username'] = user['username']
            return redirect(url_for('mainpage'))
        else:
            flash('Invalid email or password', 'danger')

        cursor.close()
        conn.close()
    
    return render_template('login.html')

# Logout Route
@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('username', None)
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

# # Route for handling document upload with bolsa_id
@app.route('/upload_document_bolsa/<int:bolsa_id>', methods=['POST'])
def upload_document_bolsa(bolsa_id):
    upload_folder = 'static/uploads'
    file = request.files.get('document')

    if not file or file.filename == '':
        flash('No selected file')
        return redirect(url_for('metadatapage'))  # Redirect on failure

    if file:
        filename = file.filename  # Ensure safe filename
        print(filename)
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)  # Save file to the uploads directory

        # Check if the file already exists for this bolsa_id
        connection = create_connection()
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM listas WHERE bolsa_id = %s AND file_name = %s", (bolsa_id, filename))
        exists = cursor.fetchone()[0]

        if exists > 0:
            cursor.close()
            connection.close()
            flash('File already exists')  
            return redirect(url_for('metadatapage'))  

        # Insert the filename and bolsa_id into the documents table
        insert_query = """
        INSERT INTO listas (bolsa_id, file_name) 
        VALUES (%s, %s)
        """
        cursor.execute(insert_query, (bolsa_id, filename))
        connection.commit()
        cursor.close()
        connection.close()

        return redirect(url_for('metadatapage'))  # Redirect to the bolsa page

@app.route('/download/<file_name>', methods=['GET'])
def download_file(file_name):
    # Define the directory where files are stored
    upload_folder = '/Users/rafaelpereira/Desktop/projeto_Bolsas/static/uploads'
    
    try:
        # Serve the file from the uploads directory
        return send_from_directory(upload_folder, file_name, as_attachment=True, mimetype='text/plain')
    except FileNotFoundError:
        abort(404, description="File not found")

@app.route('/minhaconta', methods=['GET', 'POST'])
def minhaconta():
    # Check if the user is logged in
    if not is_logged_in():
        flash("Please log in to access your account.", "warning")
        return redirect(url_for('login'))
    
    # Assuming user details are stored in session for simplicity
    user = {
        "username": session['username'],
        "email": session['email']  # You can retrieve more details as needed
    }
    
    return render_template('minhaconta.html', user=user)

@app.route('/user_profile/<int:user_id>', methods=['GET', 'POST'])
def user_profile(user_id):
    user_info = user_infos(user_id)  
    colocados = get_colocados_by_user_id(user_id)

    if request.method == 'POST':
        additional_info = request.form.get('additional-info')
        
        if additional_info is not None:
            update_additional_info(user_id, additional_info)
            flash('Informações adicionais atualizadas com sucesso!', 'success')
            return redirect(url_for('user_profile', user_id=user_id))

    return render_template('user_profile.html', user_info=user_info, colocados=colocados)



@app.route('/upload_document/<int:user_id>', methods=['POST'])
def upload_document(user_id):
    # Ensure the upload folder exists
    upload_folder = 'static/uploads'
    if not os.path.exists(upload_folder):
        os.makedirs(upload_folder)

    # Handle the file upload
    if 'documento' not in request.files:
        flash('No file part')
        return redirect(request.url)

    file = request.files['documento']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    # Save the file
    file_path = os.path.join(upload_folder, file.filename)
    file.save(file_path)

    # Optionally, you can also save the filename to the database here
    # Save the filename to the database
    connection = connect_to_database()
    cursor = connection.cursor()
    
    try:
        cursor.execute("INSERT INTO documents (user_id, file_name) VALUES (%s, %s)", (user_id, file.filename))
        connection.commit()
    except Exception as e:
        print(f"Error saving document to DB: {e}")
    finally:
        cursor.close()
        connection.close()


    return redirect(url_for('user_profile', user_id=user_id))  # Adjust this to your user profile route


@app.route('/remove_document/<int:user_id>/<path:file_name>', methods=['POST'])
def remove_document(user_id, file_name):
    connection = connect_to_database()  # Ensure this function connects to your database
    cursor = connection.cursor()

    try:
        file_path = os.path.join('static/uploads', file_name)

        # Remove the file from the filesystem
        if os.path.exists(file_path):
            os.remove(file_path)
            flash(f'Documento {file_name} removido com sucesso.')

            # Remove the document entry from the database
            cursor.execute("DELETE FROM documents WHERE file_name = %s", (file_name,))
            connection.commit()
        else:
            flash(f'Erro: O documento {file_name} não foi encontrado no sistema.')

    except Exception as e:
        print(f"Error: {e}")
        flash('Erro ao remover o documento.')

    finally:
        cursor.close()
        connection.close()

    return redirect(url_for('user_profile', user_id=user_id))  # Redirect back to user profile

@app.route('/update_account', methods=['POST'])
def update_account():
    # Check if the user is logged in
    if not is_logged_in():
        flash("Please log in to update your account.", "warning")
        return redirect(url_for('login'))
    
    # Retrieve form data
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    confirm_password = request.form['confirm_password']
    
    # Initialize hashed_password
    hashed_password = None

    # Validate and hash the password if provided
    if password:
        if password == confirm_password:
            hashed_password = generate_password_hash(password)
            #print(hashed_password)
        else:
            flash("Passwords do not match!", "danger")
            return redirect(url_for('minhaconta'))  # Redirect back to the account details page

    try:
        connection = connect_to_database()
        cursor = connection.cursor()

        # Prepare the update query
        update_query = "UPDATE Admin SET username = %s, email = %s"
        values = [username, email]

        # If a new password is being set, include it in the update
        if hashed_password:
            update_query += ", password = %s"
            values.append(hashed_password)

        # Complete the query with the WHERE clause
        update_query += " WHERE id = %s"
        values.append(session['user_id'])  # Ensure user ID is in session

        cursor.execute(update_query, tuple(values))
        connection.commit()
        
        flash("Account updated successfully!", "success")
    except Exception as e:
        flash("An error occurred while updating the account: " + str(e), "danger")
    finally:
        cursor.close()
        connection.close()
    
    return redirect(url_for('minhaconta'))  # Redirect back to the account details page

@app.route('/mainpage')
def mainpage():
    no_bolsas = total_bolsas()
    no_escolas = total_escolas()
    no_users = total_users()
    no_colocados = total_colocados()
    all_schools = get_all_escola_names()
    colocados = get_colocados()
    vagas = get_vagas_per_bolsa()

    for vaga in vagas:
        vaga['total_vagas'] = int(vaga['total_vagas'] or 0)  


    colocados_per_escola = {escola: 0 for escola in all_schools}
    total_aceite = 0
    total_negado = 0

    for colocado in colocados:
        escola_nome = colocado['escola_nome']
        estado = colocado['estado']
        if escola_nome in colocados_per_escola:
            if estado == 'aceite':
                colocados_per_escola[escola_nome] += 1
                total_aceite += 1
            elif estado == 'negado':
                total_negado += 1

    return render_template(
        'main.html',
        no_bolsas=no_bolsas,
        no_escolas=no_escolas,
        no_users=no_users,
        no_colocados=no_colocados,
        colocados_per_escola=colocados_per_escola,
        total_aceite=total_aceite,
        total_negado=total_negado,
        vagas=vagas  # Pass fixed data
    )
@app.route('/import_users_data')
def import_users():
    
    return render_template('import_user_data.html')


@app.route('/process_csv', methods=['POST'])
def process_csv():
    if 'file' not in request.files:
        flash('No file part')
        return redirect(request.url)

    file = request.files['file']
    if file.filename == '':
        flash('No selected file')
        return redirect(request.url)

    if file and allowed_file(file.filename):
        filename = file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)  # Save the file to the server

        # Process the file
        process_csv_and_update_db(file_path)

        flash('CSV file successfully processed and database updated!')
        return redirect(url_for('import_users'))

    flash('Invalid file type. Please upload a .csv file.')
    return redirect(url_for('import_users'))

@app.route('/limpar_estados', methods=['POST'])
def limpar_estados():
    connection = create_connection()
    try:
        with connection.cursor() as cursor:
            # Update all rows in Users to set estado to 'livre'
            update_query = "UPDATE Users SET estado = 'livre'"
            cursor.execute(update_query)
            
            # Delete all rows from colocados
            delete_query = "DELETE FROM colocados"
            cursor.execute(delete_query)
            
            # Delete all rows from colocados
            delete_query2 = "DELETE FROM vagas_per_bolsa"
            cursor.execute(delete_query2)
            
            # Delete all rows from colocados
            delete_query3 = "DELETE FROM documents"
            cursor.execute(delete_query3)
            
            # Commit the changes
            connection.commit()
    except Exception as e:
        print(f"Error updating estados or deleting colocados: {e}")
    finally:
        connection.close()
    
    return redirect(url_for('metadatapage'))  # Redirect back to the metadata page

# Route for selection page (same layout as main page)
@app.route('/selectionpage', methods=['GET', 'POST'])
def selection_page():
    # This is an example, adjust the query to your actual table structure
    connection = create_connection()
    cursor = connection.cursor()

    # Fetch the list of schools from the database
    cursor.execute("SELECT id, nome FROM Escola")
    escolas = cursor.fetchall()

    # Fetch the list of bolsas from the database (if you have it)
    cursor.execute("SELECT id, nome FROM Bolsa")
    bolsas = cursor.fetchall()

    connection.close()

    return render_template('selection.html', escolas=escolas, bolsas=bolsas)

@app.route('/submit_selection', methods=['POST'])
def submit_selection():
    bolsas_ids = []  # Can select multiple bolsas
    date_today = date.today()

    # Get the total number of candidates (numero de candidatos necessários)
    total_vagas = int(request.form['numero'])

    # Get the list of schools with vacancies
    escolas_data = request.form.getlist('escolas[]')
    vagas_per_escola = {}
    

    
    for escola_data in escolas_data:
        # Assuming escola_data is formatted as 'escola_nome:vagas_normais:vagas_deficiencia:bolsa_id:distribuicao'
        escola_nome, vagas_normais, vagas_deficiencia, bolsa_id, distribuicao = escola_data.split(':')
        vagas_per_escola[escola_nome] = {
            'vagas_normais': int(vagas_normais),
            'vagas_deficiencia_obrigatorias': int(vagas_deficiencia),
            'bolsa_id': bolsa_id,
            'distribuicao': distribuicao
        }
        
        # Calculate total vagas
        total_vagas = int(vagas_normais) + int(vagas_deficiencia)
        
        # Cursor for executing SQL queries
        connection = connect_to_database()
        cursor = connection.cursor()
        
        
        # Prepare the INSERT statement
        sql = """
        INSERT INTO vagas_per_bolsa (bolsa_id,total_vagas)
        VALUES (%s,%s);
        """
        
        # Execute the query
        cursor.execute(sql, (bolsa_id, total_vagas))
        
        # Commit after each insert (or batch commit later)
        connection.commit()
        cursor.close()
        connection.close()

        bolsas_ids.append(bolsa_id)

    contrato_tipo = request.form['contrato_id']
    

    initial_vagas_per_escola = {escola_nome: {
        'vagas_normais': vagas_info['vagas_normais'],
        'vagas_deficiencia_obrigatorias': vagas_info['vagas_deficiencia_obrigatorias']
    } for escola_nome, vagas_info in vagas_per_escola.items()}

    # Log the vacancies per school
    print("\nVacancies per School:")
    print(f"{'Escola':<30} {'Vagas Normais':<15} {'Vagas Deficiência':<20}")
    print("-" * 65)
    for escola_nome, vagas_info in vagas_per_escola.items():
        print(f"{escola_nome:<30} {vagas_info['vagas_normais']:<15} {vagas_info['vagas_deficiencia_obrigatorias']:<20}")

    selected_candidates = set()
    candidates_by_school = {}
    all_candidates = []

    # Aggregate all candidates across all bolsas_ids
    for bolsa_id in bolsas_ids:
        query = """
            SELECT u.id AS candidato_id, u.nome, u.nota_final, u.deficiencia, ue.escola_priority_id, ue.escola_id, e.nome AS escola_nome
            FROM Users u
            JOIN userbolsas ub ON u.id = ub.user_id
            JOIN user_escola ue ON u.id = ue.user_id
            JOIN Escola e ON ue.escola_id = e.id
            LEFT JOIN colocados c ON u.id = c.user_id
            WHERE ub.Bolsa_id = %s
            AND (
                (u.estado = 'livre')  
                OR (u.estado = 'aceite' AND c.contrato_id = 2)  
            )
            AND (
                (%s = 1 AND (ub.contrato_id = 1 OR ub.contrato_id = 3))  
                OR (%s = 2 AND (ub.contrato_id = 2 OR ub.contrato_id = 3))  
                OR (%s = 3 AND (ub.contrato_id = 1 OR ub.contrato_id = 2 OR ub.contrato_id = 3))  
            )
        """
        candidates = execute_query(query, (bolsa_id, contrato_tipo, contrato_tipo, contrato_tipo))
        all_candidates.extend(candidates)

    # Sort all candidates by nota_final (DESC) and escola_priority_id (ASC)
    all_candidates = sorted(all_candidates, key=lambda x: (-x['nota_final'], x['escola_priority_id']))

    # Allocate candidates to schools based on vacancies
    for candidate in all_candidates:
        candidato_id = candidate['candidato_id']
        candidato_nome = candidate['nome']
        candidato_nota = candidate['nota_final']
        candidato_priority = candidate['escola_priority_id']
        candidato_escola_nome = candidate['escola_nome']
        candidato_deficiencia = candidate['deficiencia']

        if candidato_escola_nome in vagas_per_escola and candidato_id not in selected_candidates:
            vagas_info = vagas_per_escola[candidato_escola_nome]
            normal_vagas = vagas_info['vagas_normais']
            vagas_deficiencia_obrigatorias = vagas_info['vagas_deficiencia_obrigatorias']

            if candidato_deficiencia == 'sim' and vagas_deficiencia_obrigatorias > 0:
                candidates_by_school.setdefault(candidato_escola_nome, []).append(candidate)
                selected_candidates.add(candidato_id)
                vagas_per_escola[candidato_escola_nome]['vagas_deficiencia_obrigatorias'] -= 1
            elif normal_vagas > 0:
                candidates_by_school.setdefault(candidato_escola_nome, []).append(candidate)
                selected_candidates.add(candidato_id)
                vagas_per_escola[candidato_escola_nome]['vagas_normais'] -= 1

    save_candidates_to_file_and_db(candidates_by_school)
               

    # Update user status and insert selected candidates into 'Colocados' table
    for escola_nome, candidatos in candidates_by_school.items():
        distrib = vagas_per_escola[escola_nome]['distribuicao']
        for candidato in candidatos:
            
            update_query = """
            UPDATE Users
            SET estado = 'a aguardar resposta', distribuicao = %s
            WHERE id = %s
            """
            insert_query2 = """
                INSERT INTO colocados (user_id, bolsa_id, escola_nome, contrato_id, escola_priority_id, placement_date,estado)
                VALUES (%s, %s, %s, %s, %s, NOW(),'a aguardar resposta')
            """
            
                
            execute_update(update_query, (distrib, candidato['candidato_id']))
            execute_insert(insert_query2, (candidato['candidato_id'], bolsa_id, candidato['escola_nome'], contrato_tipo, candidato['escola_priority_id']))
            
            user_info = user_infos(candidato['candidato_id'])
            data_to_send = {
                'NIF': user_info.get("NIF", None),
                'Bolsa_id': bolsa_id,
                'Escola_nome': candidato['escola_nome'],
                'Data_colocacao': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Current date and time in the correct format
                'Estado': 'A aguardar resposta'
            }

            api_url = 'http://127.0.0.1:8081/colocados'  
            try:
                response = requests.post(api_url, json=data_to_send)
                if response.status_code == 200:
                    print("Data successfully sent to /colocados")
                else:
                    print(f"Error sending data: {response.status_code}")
            except requests.exceptions.RequestException as e:
                print(f"Error: {e}")

    return render_template('resultados.html', 
                           candidates_by_school=candidates_by_school, 
                           vagas_per_escola=vagas_per_escola,  
                           initial_vagas_per_escola=initial_vagas_per_escola,
                           date_today=date_today, 
                           contrato_tipo=contrato_tipo, 
                           total_vagas=total_vagas)

@app.route('/send-email', methods=['POST'])
def send_email_route():
    data = request.json
    recipient = data.get('email')
    escola = data.get('escola')
    sgc = data.get('sgc')
    message = data.get('message')

    if recipient and message:
        # Convert new lines to <br> for HTML email
        html_message = message.replace('\n', '<br>')
        
        # Call the send_email_on_selection function with HTML message
        response = send_email_on_selection(sgc, [recipient], html_message)
        
        # Ensure response is not None
        if response:
            return jsonify(response[0]), response[1]
        else:
            return jsonify({"status": "error", "message": "Failed to send email"}), 500
    else:
        return jsonify({"status": "error", "message": "Missing data"}), 400
    
@app.route('/consulta')
def metadatapage():
    
    scores = get_all_user_scores()  # Fetch paginated results
    total_count = get_total_user_count()  # Count total results
    return render_template('consulta.html', scores=scores,  total_count=total_count)

@app.route('/view_escolas/<int:user_id>/<int:bolsa_id>')
def fetch_escolas(user_id, bolsa_id):
    # print("User ID:", user_id)
    escolas = get_escolas_by_bolsa(user_id, bolsa_id)
    # print("Fetched escolas:", escolas)  # This should show the fetched results
    return jsonify(escolas)  # Return JSON response

@app.route('/view_escolas/<int:user_id>')
def fetch_escolas_user(user_id):
    escolas = get_escolas_user(user_id)
    return jsonify(escolas)  # Return JSON response

@app.route('/get_escolas/<int:bolsa_id>', methods=['GET'])
def get_escolas(bolsa_id):
    escolas = get_escola_names_by_bolsa(bolsa_id)  # Get associated escolas for the bolsa_id
    
    return {'escolas': escolas} 

@app.route('/update_status', methods=['POST'])
def update_status():
    user_logged = session['username']
    data = request.json
    user_id = data['user_id']
    new_status = data['new_status']
    
    conn = create_connection()  # Assuming this creates a MySQL connection
    cursor = conn.cursor()

    try:
        # Update the user's status in the Users table
        update_query = "UPDATE Users SET estado = %s WHERE id = %s"
        cursor.execute(update_query, (new_status, user_id))
        conn.commit()

        # Fetch the last row for this user from colocados
        fetch_query = """
        SELECT 
            id, user_id, bolsa_id, escola_nome, contrato_id, escola_priority_id, placement_date, estado
        FROM colocados
        WHERE user_id = %s
        ORDER BY placement_date DESC LIMIT 1
        """
        cursor.execute(fetch_query, (user_id,))
        last_row = cursor.fetchone()

        if last_row:
            # Insert a new row in colocados with the same data but updated estado
            insert_query = """
            INSERT INTO colocados (user_id, bolsa_id, escola_nome, contrato_id, escola_priority_id, placement_date, estado,alterado_por)
            VALUES (%s, %s, %s, %s, %s, NOW(), %s,%s)
            """
            cursor.execute(insert_query, (
                last_row[1],  # user_id
                last_row[2],  # bolsa_id
                last_row[3],  # escola_nome
                last_row[4],  # contrato_id
                last_row[5],  # escola_priority_id
                new_status,
                user_logged
            ))
            conn.commit()
        
        

        return jsonify({'success': True})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'success': False}), 500

    finally:
        cursor.close()
        conn.close()


@app.route('/Bolsas/SaoMiguel')
def bolsa_sao_miguel():
    bolsa_id = 42
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/SaoMiguel.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa São Miguel"

        # Write headers
        headers = ["Nome", "Prova de Conhecimentos", "Avaliação Curricular", "Nota Final", "Escolas e Prioridades", "Contrato ID"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_saomiguel.xlsx'
        )

    return render_template('/Bolsas/SaoMiguel.html', user_info=paginated_user_info, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Terceira', methods=['GET', 'POST'])
def bolsa_terceira():
    bolsa_id = 43
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Terceira.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Terceira"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_terceira.xlsx'
        )

    return render_template('/Bolsas/Terceira.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)


@app.route('/Bolsas/SantaMaria')
def bolsa_santa_maria():
    bolsa_id = 41
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/SantaMaria.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Santa Maria"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_santamaria.xlsx'
        )

    return render_template('/Bolsas/SantaMaria.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Faial')
def bolsa_faial():
    bolsa_id = 47
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Faial.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Faial"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_faial.xlsx'
        )

    return render_template('/Bolsas/Faial.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Pico')
def bolsa_pico():
    bolsa_id = 46
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Pico.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Pico"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_pico.xlsx'
        )

    return render_template('/Bolsas/Pico.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/SaoJorge')
def bolsa_sao_jorge():
    bolsa_id = 45
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/SaoJorge.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa São Jorge"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_sjorge.xlsx'
        )

    return render_template('/Bolsas/SaoJorge.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Graciosa')
def bolsa_graciosa():
    bolsa_id = 44
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Graciosa.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Graciosa"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_graciosa.xlsx'
        )
    return render_template('/Bolsas/Graciosa.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Flores')
def bolsa_flores():
    bolsa_id = 48
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Flores.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Flores"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_flores.xlsx'
        )

    return render_template('/Bolsas/Flores.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)

@app.route('/Bolsas/Corvo')
def bolsa_corvo():
    bolsa_id = 49
    page = request.args.get('page', 1, type=int)  # Get the page number, default to 1
    per_page = 10  # Number of users per page

    uploaded_documents = get_uploaded_documents(bolsa_id)  # Fetch uploaded documents for this bolsa_id
    user_ids = has_bolsa(bolsa_id)  # Get the list of user IDs for the bolsa

    if not user_ids:
        return render_template('/Bolsas/Corvo.html', user_info=[], escolas_bolsa=[], pagination=None, uploaded_documents=uploaded_documents)

    # Fetch all user info and sort by final grade before pagination
    user_info = get_user_info(user_ids)
    user_info_sorted = sorted(user_info, key=lambda x: x['nota_final'], reverse=True)

    # Paginate the sorted user info
    total_users = len(user_info_sorted)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_user_info = user_info_sorted[start:end]

    # Fetch escolas for the current page of users
    user_ids_paginated = [user['id'] for user in user_info]
    escolas_bolsa = get_escolas_by_users(user_ids_paginated, bolsa_id)

    # Calculate total number of pages
    total_pages = (total_users + per_page - 1) // per_page

    # Pagination control flags
    pagination = {
        'page': page,
        'total_pages': total_pages,
        'has_prev': page > 1,
        'has_next': page < total_pages
    }

    if request.args.get('download_xlsx') == 'true':
        # Generate XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Bolsa Corvo"

        # Write headers
        headers = ["Nome",  "Prova de Conhecimentos","Avaliação Curricular", "Nota Final", "Escolas e Prioridades"]
        ws.append(headers)

        # Write user data
        for user in user_info_sorted:
        # Fetch escolas and contrato_id for the current user
            user_escolas = [
                escola for escola in escolas_bolsa if escola['user_id'] == user['id']
            ]
            
            # Generate a text summary of the schools
            escolas_text = ", ".join(
                [
                    f"{escola['escola_nome']} (Prioridade:{escola['escola_priority_id']})"
                    for escola in sorted(user_escolas, key=lambda x: x['escola_priority_id'])
                ]
            )
            
            # Get unique contrato_ids for the user
            contrato_ids = ", ".join(
                str(contrato_id) for contrato_id in set(
                    escola['contrato_id'] for escola in user_escolas if escola['contrato_id'] is not None
                )
            )
            # Debugging: Print contrato_ids to ensure correctness
            #print(f"Contrato IDs for user {user['id']}: {contrato_ids}")

            # Append user data to the row
            row = [
                user['nome'],
                user['prova_de_conhecimentos'],
                user['avaliacao_curricular'],
                user['nota_final'],
                escolas_text,
                contrato_ids,  # Add contrato_id to the row
            ]
            ws.append(row)

        # Save the workbook to a BytesIO stream
        xlsx_io = io.BytesIO()
        wb.save(xlsx_io)
        xlsx_io.seek(0)

        # Send XLSX as a response
        return send_file(
            xlsx_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='bolsa_corvo.xlsx'
        )

    return render_template('/Bolsas/Corvo.html', user_info=paginated_user_info, escolas_bolsa=escolas_bolsa, pagination=pagination, uploaded_documents=uploaded_documents)




@app.route('/add_user', methods=['GET', 'POST'])
def add_user():
    if request.method == 'POST':
        try:
            # Fetch the form data
            nome = request.form['nome']
            contacto = request.form['contacto']
            deficiencia = request.form['deficiencia']
            avaliacao_curricular = request.form['avaliacao_curricular']
            prova_de_conhecimentos = request.form['prova_de_conhecimentos']
            nota_final = request.form['nota_final']
            estado = request.form['estado']
            observacoes = request.form['observacoes']
            bolsa_ids = request.form.getlist('bolsa_id[]')
            contrato_ids = [request.form.get(f'contrato_id_{bolsa_id}') for bolsa_id in bolsa_ids]

            # Handle the escolas and their priorities
            escolas_per_bolsa = []
            for bolsa_id in bolsa_ids:
                escola_ids = request.form.getlist(f'escola_id_{bolsa_id}[]')  # Schools selected for this bolsa
                for escola_id in escola_ids:
                    escola_priority_id = request.form.get(f'order_id_{bolsa_id}_{escola_id}')  # Get priority/order
                    escolas_per_bolsa.append((bolsa_id, escola_id, escola_priority_id))

            connection = create_connection()
            cursor = connection.cursor()

            # Insert into Users table
            user_query = """
            INSERT INTO Users (nome, contacto, deficiencia, avaliacao_curricular, 
                               prova_de_conhecimentos, nota_final, estado, observacoes)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(user_query, (nome, contacto, deficiencia, avaliacao_curricular, 
                                         prova_de_conhecimentos, nota_final, estado, observacoes))
            user_id = cursor.lastrowid  # Get the last inserted user ID

            # Insert into userbolsas table
            for bolsa_id, contrato_id in zip(bolsa_ids, contrato_ids):
                bolsa_query = """
                INSERT INTO userbolsas (user_id, Bolsa_id, contrato_id)
                VALUES (%s, %s, %s)
                """
                cursor.execute(bolsa_query, (user_id, bolsa_id, contrato_id))

            # Insert into user_escola table
            for bolsa_id, escola_id, escola_priority_id in escolas_per_bolsa:
                escola_query = """
                INSERT INTO user_escola (user_id, escola_id, escola_priority_id)
                VALUES (%s, %s, %s)
                """
                cursor.execute(escola_query, (user_id, escola_id, escola_priority_id))

            connection.commit()
            cursor.close()
            connection.close()

            # After successfully processing POST, redirect to another page
            return redirect(url_for('mainpage'))  # Adjust redirect as needed

        except Exception as e:
            # Handle any errors, log them, and return an error message
            print(f"Error: {e}")
            return "An error occurred during form submission.", 500

    # For GET request, render the form
    connection = create_connection()
    cursor = connection.cursor()
    
    # Fetch bolsas and escolas from the database
    cursor.execute("SELECT id, nome FROM Bolsa")
    bolsas = cursor.fetchall()
    
    escolas_per_bolsa = {}
    for bolsa in bolsas:
        bolsa_id = bolsa[0]
        cursor.execute("""
            SELECT e.id, e.nome 
            FROM Escola e
            JOIN Bolsa_Escola be ON e.id = be.escola_id
            WHERE be.bolsa_id = %s
        """, (bolsa_id,))
        escolas_per_bolsa[bolsa_id] = cursor.fetchall()  # Store escolas for this bolsa
    
    cursor.close()
    connection.close()

    # Render the form template on GET request
    return render_template('add_user.html', bolsas=bolsas, escolas_per_bolsa=escolas_per_bolsa)

# # Route to receive and store data
# @app.route('/receive_data', methods=['POST'])
# def receive_data():
#     data = request.get_json()
    
#     if data:
#         connection = create_connection()
#         cursor = connection.cursor()

#         insert_query = """
#         INSERT INTO received_data (data) VALUES (%s)
#         """
        
#         try:
#             cursor.execute(insert_query, (str(data),))
#             connection.commit()
#             return jsonify({"message": "Data received and stored successfully"}), 201
#         except Error as e:
#             print(f"Error occurred: {e}")
#             return jsonify({"error": "Failed to store data"}), 500
#         finally:
#             cursor.close()
#             connection.close()
#     return jsonify({"error": "No data received"}), 400

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=8080)